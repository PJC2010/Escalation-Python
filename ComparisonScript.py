import pandas as pd
# import matplotlib as plt # Keep commented if not strictly needed
import seaborn as sns # Keep commented if not strictly needed
import os
import matplotlib.pyplot as plt # Needed for create_practice_visualization
import re # Import regular expressions module

from datetime import datetime, timedelta
from pathlib import Path
import warnings
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage # Import specifically

warnings.simplefilter(action='ignore', category=UserWarning)
warnings.simplefilter(action='ignore', category=FutureWarning)

class WorklistAnalyzer:

    def __init__(self):
        """ Initialize the WorklistAnalyzer """
        self.base_path = None
        self.output_folder = None
        self.current_date = None # Format 'MM.DD'
        self.previous_date = None # Format 'MM.DD'
        self.current_year = datetime.now().year

    def set_date(self, date_str):
        """ Set current date ('MM.DD') and calculate previous date. """
        try:
            datetime.strptime(date_str, '%m.%d') # Validate format
            self.current_date = date_str
            print(f"Current week date set to: {self.current_date}")

            current_dt_obj_for_calc = datetime.strptime(f"{self.current_year}.{date_str}", '%Y.%m.%d')
            previous_dt_obj = current_dt_obj_for_calc - timedelta(days=7)
            self.previous_date = previous_dt_obj.strftime('%m.%d')
            print(f"Previous week date calculated as: {self.previous_date}")
        except ValueError:
            raise ValueError("Invalid date format. Please use MM.DD format (e.g., '04.21')")

    def get_this_monday(self):
        """Get the date of the current week's Monday in mm.dd format"""
        # Note: This might not align with the manually set date's week
        today = datetime.now()
        monday = today - timedelta(days=today.weekday())
        return monday.strftime('%m.%d')

    def get_next_monday(self):
        """Get the date of the upcoming Monday relative to current date"""
        today = datetime.now()
        ref_date = datetime.strptime(f"{self.current_year}.{self.current_date}", '%Y.%m.%d') if self.current_date else today
        days_ahead = (7 - ref_date.weekday()) % 7
        if days_ahead == 0:
             days_ahead = 7 # Go to next week's Monday if today is Monday
        next_monday = ref_date + timedelta(days=days_ahead)
        return next_monday.strftime('%m.%d')

    def find_week_folder_by_date(self, date_str_mm_dd):
        """ Find a folder matching 'Week of' and the date (MM.DD or M.D)."""
        if not self.base_path or not self.base_path.exists():
            print(f"Base path does not exist or not set: {self.base_path}")
            return None
        try:
             # Use a dummy year like 2000 for parsing just MM.DD
             dt_obj = datetime.strptime(f"2000.{date_str_mm_dd}", '%Y.%m.%d')
             date_str_m_d = f"{dt_obj.month}.{dt_obj.day}"
        except ValueError:
             print(f"Warning: Could not parse date '{date_str_mm_dd}' to create M.D format.")
             date_str_m_d = None

        search_pattern_mm_dd = f"Week of {date_str_mm_dd}"
        search_pattern_m_d = f"Week of {date_str_m_d}" if date_str_m_d else None
        print(f"Searching for folders like: '{search_pattern_mm_dd}' OR '{search_pattern_m_d}'")

        matching_folders = []
        for folder in self.base_path.iterdir():
            if folder.is_dir():
                # Prioritize exact match if both might exist (e.g., Week of 10.1 vs Week of 1.10)
                if search_pattern_mm_dd in folder.name:
                    matching_folders.append({'path': folder, 'priority': 1, 'name': folder.name})
                    print(f"  Found potential match (MM.DD): {folder.name}")
                elif search_pattern_m_d and search_pattern_m_d in folder.name:
                     matching_folders.append({'path': folder, 'priority': 2, 'name': folder.name})
                     print(f"  Found potential match (M.D): {folder.name}")

        if not matching_folders:
            print(f"No folder found containing pattern: '{search_pattern_mm_dd}' or '{search_pattern_m_d}'")
            return None

        # Sort by priority (prefer MM.DD match) then by name
        matching_folders.sort(key=lambda x: (x['priority'], x['name']))

        if len(matching_folders) > 1:
            print(f"Warning: Multiple folders found matching patterns. Using the best match: {matching_folders[0]['name']}")

        found_folder = matching_folders[0]['path']
        print(f"Selected folder: {found_folder.name}")
        return found_folder

    def get_week_folder(self, date_to_find_mm_dd):
        """ Get the week folder path for a specific date ('MM.DD')"""
        if not date_to_find_mm_dd:
             print("Error: No date provided to find folder for.")
             return None
        folder_path = self.find_week_folder_by_date(date_to_find_mm_dd)
        if folder_path is None:
             print(f"Warning: Folder not found for date {date_to_find_mm_dd} using search patterns.")
             return None
        return folder_path

    def find_excel_files_in_folder(self, folder_path, date_str_mm_dd):
        """
        Find Excel files within the folder that match the pattern
        'MM.DD [MarketName] Med Adherence Escalations.xlsx' for the given date.
        """
        if not folder_path or not folder_path.exists():
            print(f"Folder not found or not specified: {folder_path}")
            return []

        found_files = []
        # Regex: Start, escaped date, space, any chars (market), fixed string, .xlsx or .xls, end
        pattern_str = rf"^{re.escape(date_str_mm_dd)} .+ Med Adherence Escalations\.(xlsx|xls)$"
        try:
            file_pattern = re.compile(pattern_str, re.IGNORECASE)
            print(f"Searching for files in '{folder_path.name}' matching pattern: '{pattern_str}'")
        except re.error as e:
             print(f"Error compiling regex pattern '{pattern_str}': {e}. Cannot search for files.")
             return []

        for item in folder_path.iterdir():
            if item.is_file() and not item.name.startswith('~'):
                if file_pattern.match(item.name):
                    found_files.append(item)
                    print(f"  Found matching file: {item.name}")

        if not found_files:
            print(f"No files found in '{folder_path.name}' matching the pattern for date {date_str_mm_dd}.")
            return []

        print(f"Found {len(found_files)} matching file(s) for {date_str_mm_dd} in: {folder_path.name}")
        return found_files

    def read_excel_safely(self, file_path):
        """Safely read Excel file with multiple fallback options"""
        try:
            df = pd.read_excel(file_path, engine='openpyxl', data_only=True)
            return df
        except Exception as e1:
            try:
                df = pd.read_excel(file_path, engine='openpyxl')
                return df
            except Exception as e2:
                if file_path.suffix.lower() == '.xls':
                    try:
                        df = pd.read_excel(file_path, engine='xlrd')
                        return df
                    except Exception as e3:
                         print(f"Failed to read {file_path.name} with all methods:")
                         print(f"  openpyxl(data_only): {str(e1)}")
                         print(f"  openpyxl: {str(e2)}")
                         print(f"  xlrd: {str(e3)}")
                         return None
                else:
                     print(f"Failed to read {file_path.name} with openpyxl:")
                     print(f"  data_only=True: {str(e1)}")
                     print(f"  data_only=False: {str(e2)}")
                     return None

    def _process_single_week_data(self, date_str_mm_dd, is_comparison_data=False):
        """ Processes worklist data for a single week ('MM.DD'). """
        print(f"\n--- Processing data for week of: {date_str_mm_dd} ---")
        folder_path = self.get_week_folder(date_str_mm_dd)
        if folder_path is None or not folder_path.exists():
            print(f"No valid week folder found for {date_str_mm_dd}.")
            return {}

        excel_files = self.find_excel_files_in_folder(folder_path, date_str_mm_dd)
        if not excel_files:
             print(f"No Excel files found to process for week {date_str_mm_dd}.")
             return {}

        if is_comparison_data:
            # Minimal columns needed for WoW comparison
            desired_columns = [
                'PayerMemberId', 'MarketCode', 'PracticeName', 'PCP',
                 'Escalation Path', 'Escalation Resolution', 'Gap Completed',
                 'PatientName' # Added PatientName for context in WoW lists
            ]
            processing_type = "WoW comparison (minimal columns)"
        else:
            # Full columns for main analysis
            desired_columns = [
                 'LastImpactableDate','PatientName','DateOfBirth','PracticeName','PCP',
                 'Rx Status','Call Disposition','QS Notes','Current Barrier','Action',
                 'Escalation Path','Escalation Timeframe','Escalation Deadline',
                 'Escalation Resolution','PayerCode','MarketCode','PayerMemberId',
                 'PatientPhoneNumber','PatientAddress','DataAsOfDate','EMR ID','United Flag',
                 'MedAdherenceMeasureCode','NDCDesc','Impact Category','Gap Priority',
                 'PDCNbr','ADRNbr','DaysMissedNbr','Total Fills Column?', # Check actual name if causing issues
                 'Initial Fill Date','LastFillDate','NextFillDate','DrugDispensedQuantityNbr',
                 'DrugDispensedDaysSupplyNbr','Last Activity Date','Task Status', # Check actual name
                 'OneFillCode','PrescriberNPI','PrescribingName','Prescriber Phone Number',
                 'PharmacyStoreName','PharmacyCommunicationNumberText'
             ]
            processing_type = "main analysis (full columns)"

        print(f"Processing type: {processing_type}")

        date_columns = [
            'LastImpactableDate','DateOfBirth','LastFillDate','NextFillDate',
            'Initial Fill Date','Last Activity Date','DataAsOfDate',
            'Escalation Timeframe','Escalation Deadline'
        ]

        market_dfs = {}
        total_records_processed = 0
        total_escalations_found = 0

        for file_path in excel_files:
            print(f"\nProcessing file: {file_path.name}")
            try:
                df_full_file = self.read_excel_safely(file_path)
                if df_full_file is None or df_full_file.empty:
                     print(f"  File {file_path.name} is empty or could not be read, skipping.")
                     continue

                if isinstance(df_full_file, dict): # Handle multiple sheets if necessary
                    # Basic handling: use the first sheet. Adapt if needed.
                    sheet_name = list(df_full_file.keys())[0]
                    df_sheet = df_full_file[sheet_name]
                    print(f"  Reading first sheet: '{sheet_name}' (multiple sheets found)")
                    if df_sheet.empty: continue
                else: df_sheet = df_full_file

                total_records_processed += len(df_sheet)
                df_sheet.columns = [str(col).strip() for col in df_sheet.columns]

                # --- Data Cleaning and Selection ---
                column_mapping = {}
                for desired_col in desired_columns:
                     matches = [col for col in df_sheet.columns if str(col).lower() == desired_col.lower()]
                     if matches: column_mapping[desired_col] = matches[0] # Map desired name to actual name

                required_for_processing = ['Escalation Path', 'MarketCode', 'PayerMemberId']
                missing_required = [col for col in required_for_processing if col not in column_mapping]
                if missing_required:
                    print(f"  File '{file_path.name}' missing essential columns: {missing_required}. Skipping.")
                    continue

                available_desired_cols = list(column_mapping.keys())
                df_selected = df_sheet[[column_mapping[col] for col in available_desired_cols]].copy()
                df_selected.columns = available_desired_cols # Standardize column names

                # Format date columns only if doing full processing
                if not is_comparison_data:
                     for col in date_columns:
                         if col in df_selected.columns:
                             try:
                                 df_selected[col] = pd.to_datetime(df_selected[col], errors='coerce').dt.strftime('%m/%d/%Y')
                             except Exception as date_e:
                                 print(f"    Warning: Could not format date column '{col}': {str(date_e)}")

                # Filter for relevant escalation paths
                escalation_col_name = 'Escalation Path' # Standardized name
                filtered_df = df_selected[df_selected[escalation_col_name].isin([
                    'Market/PHO Escalation',
                    'Practice Escalation'
                ])].copy()

                if filtered_df.empty:
                    print(f"  No relevant escalations found in file '{file_path.name}'.")
                    continue

                total_escalations_found += len(filtered_df)
                print(f"  Found {len(filtered_df)} relevant escalations.")

                # Group by MarketCode (using standardized name)
                market_col_name = 'MarketCode'
                for market_code in filtered_df[market_col_name].dropna().unique():
                    market_df = filtered_df[filtered_df[market_col_name] == market_code].copy()
                    market_code_str = str(market_code).strip() # Clean market code

                    if market_code_str in market_dfs:
                        market_dfs[market_code_str] = pd.concat([market_dfs[market_code_str], market_df], ignore_index=True)
                    else:
                        market_dfs[market_code_str] = market_df

            except Exception as file_e:
                print(f"Error processing file {file_path.name}: {str(file_e)}")
                # import traceback # Uncomment for detailed trace
                # print(traceback.format_exc()) # Uncomment for detailed trace

        print(f"\n--- Finished processing for week {date_str_mm_dd} ---")
        print(f"Total records scanned across files: {total_records_processed}")
        print(f"Total relevant escalations collected: {total_escalations_found}")
        print(f"Data collected for markets: {list(market_dfs.keys())}")

        return market_dfs

    def process_worklists(self):
        """ Wrapper to process the main worklist data for the current set date. """
        if not self.current_date:
             print("Error: Current date not set. Cannot process worklists.")
             return {}
        return self._process_single_week_data(self.current_date, is_comparison_data=False)

    def _get_previous_week_comparison_data(self):
        """ Wrapper to get the minimal comparison data for the previous week. """
        if not self.previous_date:
             print("Error: Previous date not calculated or set. Cannot get comparison data.")
             return {}
        return self._process_single_week_data(self.previous_date, is_comparison_data=True)

    def create_pivot_tables(self, df):
        """Create various pivot tables for analysis using standard column names"""
        pivots = {}
        if df.empty:
             print("Input DataFrame for pivot tables is empty.")
             return pivots

        # Use standardized column names now
        practice_col = 'PracticeName'
        provider_col = 'PCP'
        escalation_col = 'Escalation Path'
        member_id_col = 'PayerMemberId'

        if not all(col in df.columns for col in [practice_col, provider_col, escalation_col, member_id_col]):
             print("Warning: One or more essential columns for pivot tables are missing.")
             try:
                summary_data = self._create_summary_data(df) # Helper handles missing cols
                pivots['Summary'] = pd.DataFrame(summary_data)
             except Exception as e:
                 print(f"Could not create even summary pivot: {e}")
             return pivots

        try:
            pivots['Practice_Escalations'] = pd.pivot_table(
                df, index=practice_col, columns=escalation_col, values=member_id_col,
                aggfunc='count', fill_value=0, margins=True, margins_name='Total'
            ).sort_values('Total', ascending=False)

            df[provider_col] = df[provider_col].fillna('Unknown Provider')
            pivots['Provider_Escalations'] = pd.pivot_table(
                df, index=provider_col, columns=escalation_col, values=member_id_col,
                aggfunc='count', fill_value=0, margins=True, margins_name='Total'
            ).sort_values('Total', ascending=False)

            summary_data = self._create_summary_data(df)
            pivots['Summary'] = pd.DataFrame(summary_data)

        except Exception as e:
            print(f"Error creating pivot tables: {str(e)}")
        return pivots

    def _create_summary_data(self, df):
        """Helper function to create summary data dictionary using standard names."""
        practice_col = 'PracticeName'
        provider_col = 'PCP'
        escalation_col = 'Escalation Path'

        total_escalations = len(df)
        market_pho = len(df[df[escalation_col] == 'Market/PHO Escalation']) if escalation_col in df.columns else 0
        practice_esc = len(df[df[escalation_col] == 'Practice Escalation']) if escalation_col in df.columns else 0
        unique_practices = df[practice_col].nunique() if practice_col in df.columns else 0
        unique_providers = df[provider_col].nunique() if provider_col in df.columns else 0

        return {
                'Metric': ['Total Escalations','Market/PHO Escalations','Practice Escalations','Unique Practices','Unique Providers','Report Generated'],
                'Value': [total_escalations, market_pho, practice_esc, unique_practices, unique_providers, datetime.now().strftime('%Y-%m-%d %H:%M')]
            }

    def create_practice_visualization(self, pivot_df, market_code, output_filepath_xlsx):
        """Creates and saves a stacked bar chart PNG."""
        if pivot_df.empty or 'Total' not in pivot_df.index:
             print(f"Skipping visualization for {market_code}: Pivot data invalid or empty.")
             return None

        try:
            plt.style.use('seaborn-v0_8-whitegrid')
            viz_df = pivot_df.drop('Total', axis=0)
            if 'Total' in viz_df.columns: viz_df = viz_df.drop('Total', axis=1)
            if viz_df.empty: return None

            num_practices = len(viz_df)
            fig_height = max(6, num_practices * 0.35)
            fig, ax = plt.subplots(figsize=(10, fig_height))
            viz_df.plot(kind='barh', stacked=True, ax=ax, colormap='viridis')

            ax.set_title(f'{market_code} Escalations by Practice', pad=15, fontsize=12, weight='bold')
            ax.set_xlabel('# of Escalations', fontsize=10); ax.set_ylabel('')
            ax.tick_params(axis='y', labelsize=8); ax.tick_params(axis='x', labelsize=9)

            for container in ax.containers:
                labels = [f'{int(v)}' if v > 0 else '' for v in container.datavalues]
                ax.bar_label(container, labels=labels, label_type='center', fontsize=7, color='white', weight='bold')

            ax.invert_yaxis()
            ax.legend(title='Escalation Type', bbox_to_anchor=(1.02, 1), loc='upper left', fontsize=9, title_fontsize=10)
            plt.tight_layout(rect=[0, 0, 0.9, 1])

            # Construct PNG filename based on XLSX filename stem
            png_filename = output_filepath_xlsx.stem + "_Practice_Chart.png"
            plot_filepath = output_filepath_xlsx.parent / png_filename
            plt.savefig(plot_filepath, dpi=300, bbox_inches='tight')
            plt.close(fig)
            print(f"Created visualization: {plot_filepath.name}")
            return plot_filepath

        except Exception as e:
            print(f"Error creating visualization for {market_code}: {str(e)}")
            plt.close() # Ensure plot closed on error
            return None

    def _insert_image_to_excel(self, xlsx_path, img_path, sheet_name='Practice Chart', cell='B2'):
        """ Helper to insert image into an existing excel file """
        if not img_path or not img_path.exists():
             print(f"Cannot insert image: image file not found at {img_path}")
             return
        if not xlsx_path or not xlsx_path.exists():
             print(f"Cannot insert image: excel file not found at {xlsx_path}")
             return
        try:
            workbook = load_workbook(xlsx_path)
            if sheet_name not in workbook.sheetnames:
                 workbook.create_sheet(sheet_name)
                 print(f"Created sheet '{sheet_name}' for chart.")
            worksheet = workbook[sheet_name]
            img_obj = ExcelImage(img_path)
            # Optional: Adjust image size if needed
            # img_obj.width *= 0.75
            # img_obj.height *= 0.75
            worksheet.add_image(img_obj, cell)
            workbook.save(xlsx_path)
            print(f"Inserted image {img_path.name} into sheet '{sheet_name}' at cell {cell}.")
        except ImportError:
             print("Warning: Pillow not installed? Cannot insert image. `pip install Pillow`")
        except Exception as e:
             print(f"Error inserting image {img_path.name} into sheet '{sheet_name}': {str(e)}")


    def create_market_files(self, current_market_dfs):
        """
        Create separate Excel files for each market including raw data, pivots,
        and the new Week-over-Week comparison sheets.
        Uses output filename format: MM.DD [MarketName] Med Adherence Escalations.xlsx
        """
        if not current_market_dfs:
            print("No current week data available to create files.")
            return

        print("\n--- Preparing Week-over-Week Comparison Data ---")
        previous_market_dfs_comp = self._get_previous_week_comparison_data()
        current_market_dfs_comp = self._process_single_week_data(self.current_date, is_comparison_data=True)

        if not previous_market_dfs_comp: print("Warning: No previous week data found for comparison.")
        if not current_market_dfs_comp: print("Warning: Could not process current week data for comparison.")

        file_date_prefix = self.current_date # Use MM.DD format
        print(f"\n--- Generating Market Reports for Week {file_date_prefix} ---")

        all_market_codes = set(current_market_dfs.keys()) | set(previous_market_dfs_comp.keys())

        for market_code in all_market_codes:
            print(f"\nProcessing market: {market_code}")

            current_df_full = current_market_dfs.get(market_code, pd.DataFrame())
            current_df_comp = current_market_dfs_comp.get(market_code, pd.DataFrame())
            previous_df_comp = previous_market_dfs_comp.get(market_code, pd.DataFrame())

            if current_df_full.empty and current_df_comp.empty:
                 print(f"No current week data found for market {market_code}. Skipping file creation.")
                 continue

            # Output Filename Format
            filename = f"{file_date_prefix} {market_code} Med Adherence Escalations.xlsx"
            file_path = self.output_folder / filename
            print(f"Output file will be: {filename}")

            img_filepath_to_insert = None # Reset for each market

            try:
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    # --- 1. Write Main Analysis Tabs ---
                    if not current_df_full.empty:
                        data_sheet_name = f"{market_code} Data"
                        print(f"- Writing '{data_sheet_name}' sheet ({len(current_df_full)} records)...")
                        current_df_full.to_excel(writer, sheet_name=data_sheet_name, index=False)
                        # Autofit columns for data sheet
                        worksheet = writer.sheets[data_sheet_name]
                        for idx, column in enumerate(current_df_full.columns):
                             col_letter = get_column_letter(idx + 1)
                             try: max_len = max(current_df_full[column].astype(str).map(len).max(), len(str(column)))
                             except: max_len = len(str(column)) # Fallback
                             adjusted_width = min((max_len + 2) * 1.1, 60)
                             worksheet.column_dimensions[col_letter].width = adjusted_width

                        print("- Creating and writing Pivot Table sheets...")
                        pivot_tables = self.create_pivot_tables(current_df_full)
                        for pivot_name, pivot_df in pivot_tables.items():
                             if not pivot_df.empty:
                                 sheet_name = pivot_name[:31]
                                 pivot_df.to_excel(writer, sheet_name=sheet_name)
                                 print(f"  - Created '{sheet_name}' sheet.")
                                 # Autofit columns for pivot sheets
                                 pivot_worksheet = writer.sheets[sheet_name]
                                 for col_idx, col_val in enumerate(pivot_df.reset_index().columns):
                                     col_letter = get_column_letter(col_idx + 1)
                                     try: max_len = max(pivot_df.reset_index()[col_val].astype(str).map(len).max(), len(str(col_val)))
                                     except: max_len = len(str(col_val))
                                     adjusted_width = min((max_len + 2) * 1.1, 50)
                                     pivot_worksheet.column_dimensions[col_letter].width = adjusted_width
                             else: print(f"  - Pivot table '{pivot_name}' was empty.")

                         # Create visualization PNG (don't insert yet)
                        practice_pivot = pivot_tables.get('Practice_Escalations')
                        if practice_pivot is not None and not practice_pivot.empty:
                              print("- Creating Practice Escalation visualization PNG...")
                              img_filepath_to_insert = self.create_practice_visualization(practice_pivot, market_code, file_path)
                              if not img_filepath_to_insert: print("  - Visualization PNG creation failed.")
                        else: print("- Skipping Practice Escalation visualization (no data).")
                    else: print("- No current week data to write main analysis tabs.")

                    # --- 2. Perform WoW Comparison and Write Tabs ---
                    print("- Performing Week-over-Week comparison...")
                    new_members = pd.DataFrame()
                    resolved = pd.DataFrame()
                    wow_summary_dict = {'Metric': ['Current Week Escalations', 'Previous Week Escalations', 'New Escalations This Week', 'Removed Since Last Week', 'Net Change', 'Report Generated'], 'Value': [0, 0, 0, 0, 0, datetime.now().strftime('%Y-%m-%d %H:%M')]}
                    id_col = 'PayerMemberId' # Standardized name
                    current_ids = set()
                    prev_ids = set()

                    if id_col in current_df_comp.columns: current_ids = set(current_df_comp[id_col].dropna().unique())
                    else: print(f"  Warning: Comparison ID '{id_col}' not in current data for {market_code}.")
                    if id_col in previous_df_comp.columns: prev_ids = set(previous_df_comp[id_col].dropna().unique())
                    else: print(f"  Warning: Comparison ID '{id_col}' not in previous data for {market_code}.")

                    if current_ids or prev_ids: # Only compare if we have some IDs
                         new_ids = current_ids - prev_ids
                         resolved_ids = prev_ids - current_ids

                         if new_ids and id_col in current_df_comp.columns:
                              new_members = current_df_comp[current_df_comp[id_col].isin(new_ids)].drop_duplicates(subset=[id_col]).reset_index(drop=True)
                              print(f"  - Identified {len(new_members)} new escalations.")
                         else: print("  - No new escalations identified.")

                         if resolved_ids and id_col in previous_df_comp.columns:
                              resolved = previous_df_comp[previous_df_comp[id_col].isin(resolved_ids)].drop_duplicates(subset=[id_col]).reset_index(drop=True)
                              print(f"  - Identified {len(resolved)} removed escalations.")
                         else: print("  - No escalations removed since last week.")
                    else:
                         print("  - Skipping WoW comparison logic due to missing ID columns in data.")


                    wow_summary_dict['Value'][0] = len(current_ids)
                    wow_summary_dict['Value'][1] = len(prev_ids)
                    wow_summary_dict['Value'][2] = len(new_members)
                    wow_summary_dict['Value'][3] = len(resolved)
                    wow_summary_dict['Value'][4] = wow_summary_dict['Value'][0] - wow_summary_dict['Value'][1]
                    wow_summary_df = pd.DataFrame(wow_summary_dict)

                    print("- Writing Week-over-Week comparison sheets...")
                    wow_summary_df.to_excel(writer, sheet_name='WoW Summary', index=False)
                    # Provide default columns if dataframes are empty for WoW sheets
                    new_cols = new_members.columns if not new_members.empty else (current_df_comp.columns if not current_df_comp.empty else ['PayerMemberId','PatientName','MarketCode','PracticeName'])
                    res_cols = resolved.columns if not resolved.empty else (previous_df_comp.columns if not previous_df_comp.empty else new_cols)

                    pd.DataFrame(new_members, columns=new_cols).to_excel(writer, sheet_name='New This Week', index=False)
                    pd.DataFrame(resolved, columns=res_cols).to_excel(writer, sheet_name='Previous Week Only', index=False)


                    # Auto-fit WoW sheets
                    for sheet_name in ['WoW Summary', 'New This Week', 'Previous Week Only']:
                        if sheet_name in writer.sheets:
                            ws = writer.sheets[sheet_name]
                            df_to_size = new_members if sheet_name == 'New This Week' else resolved if sheet_name == 'Previous Week Only' else wow_summary_df
                            if df_to_size is not None and not df_to_size.empty:
                                for idx, col in enumerate(df_to_size.columns):
                                     col_letter = get_column_letter(idx + 1)
                                     try: max_len = max(df_to_size[col].astype(str).map(len).max(), len(str(col)))
                                     except: max_len = len(str(col))
                                     adjusted_width = min((max_len + 2) * 1.1, 50)
                                     ws.column_dimensions[col_letter].width = adjusted_width

                # --- Insert Image (after closing ExcelWriter) ---
                if img_filepath_to_insert and img_filepath_to_insert.exists():
                    print(f"- Attempting to insert image {img_filepath_to_insert.name} into {file_path.name}...")
                    self._insert_image_to_excel(file_path, img_filepath_to_insert, sheet_name='Practice Chart', cell='B2')
                elif img_filepath_to_insert:
                     print(f"- Image file not found, skipping insertion: {img_filepath_to_insert}")


                print(f"\nSuccessfully created report: {file_path.name}")

            except Exception as e:
                print(f"\nError creating file for market {market_code}: {str(e)}")
                import traceback
                print(traceback.format_exc())


def main():
    
    BASE_PATH = r"C:/Users/pcastillo/OneDrive - VillageMD/Documents - VMD- Quality Leadership- PHI/Data Updates/MedAdhData Dropzone/"
    OUTPUT_FOLDER = r"C:/Users/pcastillo/OneDrive - VillageMD/Desktop/Escalation Python/"
    # OUTPUT_FOLDER = r"C:/Users/pcastillo/OneDrive - VillageMD/Desktop/Escalation Python Output/" # Test output

    # *** Set the date for the CURRENT week's report here (MM.DD format) ***
    # Should correspond to the date prefix in the input filenames for that week
    CURRENT_WEEK_DATE = "04.28" # <-- CHANGE THIS (e.g., "04.29" if running on April 29th for week of April 28th)

    # --- Execution ---
    try:
        print("--- Starting Worklist Analysis and WoW Comparison ---")
        start_time = datetime.now()
        analyzer = WorklistAnalyzer()
        analyzer.base_path = Path(BASE_PATH)
        analyzer.output_folder = Path(OUTPUT_FOLDER)

        analyzer.output_folder.mkdir(parents=True, exist_ok=True)
        print(f"Using Base Path: {analyzer.base_path}")
        print(f"Using Output Folder: {analyzer.output_folder}")

        analyzer.set_date(CURRENT_WEEK_DATE) # Sets current and previous dates

        # Process current week for main analysis dataframes
        current_market_data = analyzer.process_worklists()

        # Create the market files (which includes WoW comparison using helper methods)
        if current_market_data:
            analyzer.create_market_files(current_market_data)
            print("\n--- Processing complete! ---")
        else:
            print("\n--- No data found for the current week. No reports generated. ---")

        end_time = datetime.now()
        print(f"Total execution time: {end_time - start_time}")

    except ValueError as ve:
         print(f"Configuration Error: {str(ve)}")
    except FileNotFoundError as fnfe:
         print(f"Path Error: {str(fnfe)}. Please check BASE_PATH and OUTPUT_FOLDER.")
    except Exception as e:
        print(f"\n--- An unexpected error occurred during main execution ---")
        print(f"Error: {str(e)}")
        import traceback
        print(traceback.format_exc())

if __name__ == "__main__":
    main()